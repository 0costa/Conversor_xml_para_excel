from PyQt5.QtWidgets import QMainWindow, QApplication, QWidget, QGridLayout, QLabel, QFileDialog, QPushButton, QProgressBar, QButtonGroup, QMessageBox
from PyQt5.QtGui import QFont,QCursor
from PyQt5.QtCore import Qt
from openpyxl import Workbook, load_workbook
import os, untangle, re, datetime

class Excel:
    def __init__(self, dados:list):
        self.__dados = dados

    def criar_planilha(self):
        if not self.__dados:
            return
        
        cabecalho = ['Chave de acesso', 'Cod. Loja', 'Nome Loja', 'Cnpj Loja', 'Num. NFe', 'Nome Dest.', 'Cnpj Dest.', 'CFOP', 'Items', 'Valor Total', 'Emissão', 'Remessa', 'Logistica', 'Status']
        
        wb = Workbook()
        ws = wb.active

        try:
            self.__dados.insert(0, cabecalho)
            self.inserir_dados(ws)
            wb.save(f'C:/Users/{os.getlogin()}/Desktop/dados_xmls.xlsx')
            Alert('Planilha salva!')
        except:
            return
        
    def carregar_planilha(self, caminho):
        wb = load_workbook(caminho)
        ws = wb['NFe']

        try:
            self.inserir_dados(ws)
            wb.save(caminho)
            Alert('Planilha salva!')
        except:
            return
            
    def inserir_dados(self, ws):
        if len(self.__dados) == 1:
            ws.append(self.__dados)

        for row in self.__dados:
            ws.append(row)

class ReadXML:
    def __init__(self, file) -> None:
        self.doc = untangle.parse(file)

    def __bool__(self):
        with open('CFOP.txt', 'r') as cfop_list:
            cfop_list = cfop_list.read().splitlines()

            cfop = [x.prod.CFOP.cdata for x in self.doc.TNfeProc.NFe.infNFe.det]
            cfop = list(set(cfop))[0]

            if not cfop in cfop_list:
                return False
            return True

    def nfe(self):
        try:
            root= self.doc.TNfeProc.NFe

            chave_de_acesso= self.doc.TNfeProc.protNFe.infProt.chNFe.cdata
            codigo_loja=  int(root.infNFe.emit.xFant.cdata.split('-')[0].strip())
            nome_loja= re.sub(f'\s\s+', ' ', root.infNFe.emit.xNome.cdata.upper().strip())
            cnpj_loja= root.infNFe.emit.CNPJ.cdata
            num_nfe= int(root.infNFe.ide.nNF.cdata)
            nome_destinatario= re.sub(f'\s\s+', ' ', root.infNFe.dest.xNome.cdata.upper().strip())
            cnpj_destinatario= root.infNFe.dest.CNPJ.cdata
            cfop= int(list(set([x.prod.CFOP.cdata for x in self.doc.TNfeProc.NFe.infNFe.det]))[0])
            items= sum([int(x.prod.qCom.cdata.split('.')[0]) for x in self.doc.TNfeProc.NFe.infNFe.det])
            valor=  float(root.infNFe.total.ICMSTot.vNF.cdata)
            emissao= root.infNFe.ide.dhEmi.cdata.split('T')[0]
            emissao= datetime.datetime.strptime(emissao, '%Y-%m-%d').strftime('%d/%m/%Y')

            return [chave_de_acesso, codigo_loja, nome_loja, cnpj_loja, num_nfe, nome_destinatario, cnpj_destinatario, cfop, items, valor, emissao]
        except:
            return []
        
class Conversor:
    def pasta(self, path):
        dados = []

        for file in os.listdir(path):
            if not str(file).endswith('.xml'):
                continue
            
            file = os.path.join(path, file)
            xml = ReadXML(file)
            
            if not xml:
                continue
            
            dados.append(xml.nfe())

        return dados
    
    def arquivo(self, path):
        if not str(path).endswith('.xml'):
            Alert('Arquivo não é xml!')
            return
        
        xml = ReadXML(path)
        
        if not xml:
            Alert('Xml inválido!')
            return

        return xml.nfe()

class Alert(QMessageBox):
    def __init__(self, text) -> QMessageBox:
        super().__init__()
        self.setWindowFlags(Qt.Dialog | Qt.CustomizeWindowHint | Qt.WindowTitleHint)
        self.setWindowTitle('Alerta')
        self.setIcon(QMessageBox.Warning)
        self.setText(text)
        self.exec()

class Picker(QFileDialog):
    def __init__(self):
        super().__init__()

    @property
    def folder(self) -> str:
        path = self.getExistingDirectory()
        return path
    
    @property
    def file(self) -> str:
        path = self.getOpenFileName()
        return path

class Botao(QPushButton):
    def __init__(self, text):
        super().__init__()  
        self.setText(text) 
        self.setStyleSheet("""
                            QPushButton {
                                padding: 10px;
                                background-color: #1899D6;
                                border-bottom: 3px solid #116c99;
                                border-radius: 6px;
                                color: #FFFFFF;
                            }
                           
                            QPushButton:pressed{
                                border-bottom: None;
                            }
                           
                            QPushButton:disabled{
                                background-color:#80c0c0c0;
                                border: none;
                            }
                            """)
        self.setCursor(QCursor(Qt.PointingHandCursor))

class Label(QLabel):
    def __init__(self, text, font_size, **kwargs):
        super().__init__()

        self.setText(text)
        self.setFont(QFont('Poppins', font_size))

        if 'align_center' in kwargs:
            self.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        if 'padding' in kwargs:
            self.setStyleSheet(f'padding: {int(kwargs["padding"])}px;')

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.__setup()
        self.__layout()
        self.__signals()

    def __setup(self):
        self.__titulo = Label('Conversor de Xml para Excel', 16, align_center=True, padding=10)


        self.__label_local_xml = Label('Onde estão os arquivos xmls?', 12)
        self.__btn_local_xml_pasta = Botao('Em uma pasta')
        self.__btn_local_xml_arquivo = Botao('É somente um arquivo')

        self.__carregamento = QLabel()
        self.__carregamento.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.__carregamento.setFixedHeight(30)
        self.__carregamento.setStyleSheet("""
                                          QLabel{
                                            background-color: #ffffff;
                                            border: 1px solid lightgray; 
                                            border-radius: 14px;
                                            text-align: center;
                                          }
                                        """) 

        self.__btn_salvar_nova_plan = Botao('Salvar em uma nova planilha')
        self.__btn_salvar_atual_plan = Botao('Salvar em uma planilha existente')
        self.__btn_salvar_nova_plan.setHidden(True)
        self.__btn_salvar_atual_plan.setHidden(True)

        self.__btn_group = QButtonGroup()
        self.__btn_group.addButton(self.__btn_local_xml_pasta, 1)
        self.__btn_group.addButton(self.__btn_local_xml_arquivo, 2)
        self.__btn_group.addButton(self.__btn_salvar_nova_plan, 3)
        self.__btn_group.addButton(self.__btn_salvar_atual_plan, 4)

    def __layout(self):
        layout = QGridLayout()
        layout.setSpacing(15)
        layout.addWidget(self.__titulo, 0 ,0, 1, 3)

        layout.addWidget(self.__label_local_xml, 1, 0)
        layout.addWidget(self.__btn_local_xml_pasta, 1, 1)
        layout.addWidget(self.__btn_local_xml_arquivo, 1, 2)

        layout.addWidget(self.__carregamento, 2, 0, 1, 3)

        layout.addWidget(self.__btn_salvar_nova_plan, 3, 0)
        layout.addWidget(self.__btn_salvar_atual_plan, 3, 1)

        main_content_widget = QWidget()
        main_content_widget.setLayout(layout)
        main_content_widget.setContentsMargins(10,10,10,10)
        
        self.setCentralWidget(main_content_widget)
        #self.setStyleSheet("QWidget{background: #ffffff;}")

    def __signals(self):
        self.__btn_group.buttonClicked.connect(self.action)
    
    def action(self, button):
        button_id = self.__btn_group.id(button)

        if button_id == 1:
            folder_path = Picker().folder

            if not folder_path:
                return Alert('Nenhuma pasta selecionado!')
            
            self.__btn_local_xml_arquivo.setDisabled(True)
            self.__carregamento.setText('Carregando')

            self.dados_xml = Conversor().pasta(folder_path)

            self.__carregamento.setText('Concluido')
            self.mostrar_botoes_save

        if button_id == 2:
            file_path = Picker().file[0]

            if not file_path:
                return Alert('Nenhum arquivo selecionado!')
            
            self.__btn_local_xml_pasta.setDisabled(True)
            self.__carregamento.setText('Carregando')

            self.dados_xml = []
            self.dados_xml.append(Conversor().arquivo(file_path))

            self.__carregamento.setText('Concluido')
            self.mostrar_botoes_save

        if button_id == 3:
            dados = self.dados_xml

            planilha = Excel(dados)
            planilha.criar_planilha()
            self.esconder_botoes_save

        if button_id == 4:
            planilha_destino = Picker().file[0]

            if not planilha_destino or not planilha_destino.endswith('.xlsx'):
                return Alert('Nenhum arquivo selecionado ou planilha invalida!')
            
            dados = self.dados_xml

            planilha_dados = Excel(dados)
            planilha_dados.carregar_planilha(planilha_destino)

            self.esconder_botoes_save

    @property
    def esconder_botoes_save(self):
        self.__btn_salvar_nova_plan.setHidden(True)
        self.__btn_salvar_atual_plan.setHidden(True)
        self.__carregamento.clear()

        self.__btn_local_xml_pasta.setDisabled(False)
        self.__btn_local_xml_arquivo.setDisabled(False)

    @property
    def mostrar_botoes_save(self):
        self.__btn_salvar_atual_plan.setHidden(False)
        self.__btn_salvar_nova_plan.setHidden(False)

def main():
    app = QApplication([])
    window = MainWindow()
    window.show()
    app.exec()

if __name__ == "__main__":
    main()